###Import information###
import numpy as np
import sys
import os
# import xlrd
# import xlwt
# import openpyxl
import time

from openpyxl import load_workbook
from openpyxl.styles import colors, Font, Fill, NamedStyle
from openpyxl.styles import PatternFill, Border, Side, Alignment
from suds.client import Client

###Global Parameter###
ACS_INFO = "http://172.16.1.78/ftacsws/acsws.asmx?WSDL"
GPN = 0
GPV = 1
GPA = 0
CREATOR = "@~''~@"
APPID = "~'@@'~"
SOURCE = 0  # 0->update from client, 1 -> update from db
Test_db = "Test_db.xlsx"


#########
# Print Help information
# @param: None
# @return: None
########
def print_help():
    helpp = "The command format is: %s OUI Serial ModelName paramter" % sys.argv[0]
    print(helpp)
    return


def Init(oui, serial, modelname, parameter):
    global m_OUI
    global m_Serial
    global m_ModelName
    global m_parameter
    m_OUI = oui
    m_Serial = serial
    m_ModelName = modelname
    m_parameter = parameter
    # print("OUI" % m_OUI)
    # print("Serial" % m_Serial)
    # print("ModelName" % m_ModelName)
    # print("Parameter" % m_parameter)
    return


def Create_Client():
    global m_client
    m_client = Client(ACS_INFO)
    return


def Connect(oui, serial, modelname, paramter):
    return


def Send_Request():
    return


def Set_Parameter():
    return


def Get_Parameter():
    ArrayOfString = m_client.factory.create('ArrayOfString')
    ArrayOfString.string = [m_parameter]
    result = m_client.service.FTGetDeviceParameters(m_Serial, ArrayOfString, SOURCE, m_OUI, m_ModelName, GPN, GPV, GPA,
                                                    CREATOR, APPID)
    return result


if __name__ == '__main__':
    input_num = len(sys.argv)
    # Create client object
    Create_Client()
    ##Error handle
    if input_num < 4:
        print_help()
        print(client)
        quit()
    ##Initialize
    Init(sys.argv[1], sys.argv[2], sys.argv[3], sys.argv[4])
    # global m_OUI
    # global m_Serial
    # global m_ModelName
    # global m_parameter
    # m_OUI=sys.argv[1]
    # m_Serial=sys.argv[2]
    # m_ModelName=sys.argv[3]
    # m_parameter=sys.argv[4]

    # Get parameter
    result = Get_Parameter()
    if result.ErrorCode != 100:
        print("error:%s" % result.ErrorCode)
        quit()

    arr_aa = np.array(result.Params.ParamWSDL)
    # print(result)
    # print("ndim:{0}".format(arr_aa.ndim))
    # print("shape:{0}".format(arr_aa.shape))
    times = int(format(arr_aa.shape[0]))
    '''
	workbook = openpyxl.Workbook()
	'''
    arr_t = arr_aa.ndim

    if arr_t == 2:
        times = 1
    '''
	sheet=workbook.active
	workbook_1 = workbook['Sheet']
	sheet['A1']='Times'
	sheet['B1']='Date'
	sheet['C1']='Name'
	sheet['D1']='Value'
	sheet['E1']='Return'
	x=2
	'''

    for i in range(times):
        localtime = time.asctime(time.localtime(time.time()))
        if arr_t != 2:
            '''
        		sheet['A'+str(x)] = i +1
        		sheet['B'+str(x)] = localtime
        		sheet['C'+str(x)] = result.Params.ParamWSDL[i].Name
        		sheet['D'+str(x)] = result.Params.ParamWSDL[i].Value
        		sheet['E'+str(x)] = result.Params.ParamWSDL[i].ErrorCode
        		x+=1
        		'''
            print("Name:  %s" % result.Params.ParamWSDL[i].Name)
            print("Value: %s" % result.Params.ParamWSDL[i].Value)
            ret = result.Params.ParamWSDL[i].ErrorCode
        else:
            '''
        		sheet['A'+str(x)] = i +1
        		sheet['B'+str(x)] = localtime
        		sheet['C'+str(x)] = result.Params.ParamWSDL.Name
        		sheet['D'+str(x)] = result.Params.ParamWSDL.Value
        		sheet['E'+str(x)] = result.Params.ParamWSDL.ErrorCode
        		x+=1
        		'''
            print("Name:  %s" % result.Params.ParamWSDL.Name)
            print("Value: %s" % result.Params.ParamWSDL.Value)
            ret = result.Params.ParamWSDL.ErrorCode
            if ret == 100:
                print("Return OK.")
            else:
                if arr_t != 2:
                    print("ErrorCode: %s" % result.Params.ParamWSDL[i].ErrorCode)
                else:
                    print("ErrorCode: %s" % result.Params.ParamWSDL.ErrorCode)

                '''
				for col in workbook_1.columns:
    				max_length = 0
    				column = col[0].column

    				for cell in col:
        			if len(str(cell.value)) > max_length:
            			max_length = len(str(cell.value))

    				adjusted_width = (max_length + 2) * 1.2
    				#workbook_1.column_dimensions[column].width = adjusted_width
    				workbook_1.column_dimensions[column].bestFit = "true"
    
				workbook.save(Test_db)
				'''
# FTGetDeviceParameters(xs:string devicesn, ArrayOfString arraynames, xs:int SOURCE, xs:string OUI, xs:string modelname, xs:boolean names, xs:boolean values, xs:boolean attributes, xs:string CREATOR, xs:string APPID)
